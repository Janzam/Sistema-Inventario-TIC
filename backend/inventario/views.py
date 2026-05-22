import pandas as pd
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.db.models import Count, Q
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authtoken.models import Token
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from .models import Equipo, Profile, Categoria, Subcategoria, Persona, Noticia
from .serializers import (
    EquipoSerializer, UserSerializer, ProfileSerializer, 
    CategoriaSerializer, SubcategoriaSerializer, PersonaSerializer, NoticiaSerializer
)

from rest_framework.authtoken.views import ObtainAuthToken

# --- VISTA PARA LOGIN PERSONALIZADA ---
class CustomObtainAuthToken(ObtainAuthToken):
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data,
                                           context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        
        # Obtener rol de la persona vinculada
        role = 'VIEWER'
        persona_name = user.first_name
        if hasattr(user, 'persona'):
            role = user.persona.rol
            persona_name = user.persona.nombre
            
        return Response({
            'token': token.key,
            'user': {
                'id': user.pk,
                'username': user.username,
                'name': persona_name,
                'email': user.email,
                'rol': role
            }
        })

# --- VISTA PARA PERSONAS ---
class PersonaViewSet(viewsets.ModelViewSet):
    queryset = Persona.objects.all()
    serializer_class = PersonaSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['post'])
    def create_user(self, request, pk=None):
        persona = self.get_object()
        if persona.user:
            return Response({"error": "Esta persona ya tiene un usuario asociado"}, status=400)
        
        username = request.data.get('username')
        password = request.data.get('password')
        
        if not username or not password:
            return Response({"error": "Usuario y contraseña requeridos"}, status=400)
            
        if User.objects.filter(username=username).exists():
            return Response({"error": "El nombre de usuario ya existe"}, status=400)
            
        try:
            user = User.objects.create_user(
                username=username,
                password=password,
                email=persona.email or '',
                first_name=persona.nombre
            )
            user.is_active = True # Creados por admin son activos de una vez
            user.save()
            
            persona.user = user
            persona.rol = request.data.get('rol', 'VIEWER')
            persona.save()
            
            # Crear perfil
            Profile.objects.get_or_create(user=user)
            
            return Response({"message": "Usuario creado y vinculado correctamente"}, status=201)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def pending_approval(self, request):
        # Personas que tienen un usuario pero el usuario está is_active=False
        pending = Persona.objects.filter(user__is_active=False)
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve_user(self, request, pk=None):
        persona = self.get_object()
        if not persona.user:
            return Response({"error": "Esta persona no tiene un usuario asociado"}, status=400)
        
        rol = request.data.get('rol')
        if rol not in ['ADMIN', 'TECNICO', 'VIEWER']:
            return Response({"error": "Rol inválido"}, status=400)
            
        persona.rol = rol
        persona.save()
        
        persona.user.is_active = True
        persona.user.save()
        
        return Response({"message": f"Usuario {persona.user.username} aprobado como {rol}"})

# --- VISTA PARA CATEGORÍAS ---
class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def with_stats(self, request):
        user = request.user
        role = getattr(getattr(user, 'persona', None), 'rol', 'VIEWER')
        
        # Definir filtro de equipos según rol
        equipo_filter = Q()
        if role == 'VIEWER':
            persona = getattr(user, 'persona', None)
            if persona:
                equipo_filter = Q(subcategorias__equipos__usuario_asignado__iexact=persona.nombre)
            else:
                equipo_filter = Q(id__lt=0) # Forzar 0
        else:
            # Admins/Técnicos ven todo el conteo
            equipo_filter = Q(subcategorias__equipos__isnull=False)

        # Optimizamos con annotation para evitar el loop N+1
        categorias = Categoria.objects.annotate(
            total_items=Count('subcategorias__equipos', filter=equipo_filter, distinct=True)
        ).prefetch_related('subcategorias')

        result = []
        for cat in categorias:
            result.append({
                "id": cat.id,
                "nombre": cat.nombre,
                "imagen": cat.imagen,
                "color": cat.color,
                "total_equipos": cat.total_items,
                "subcategorias": SubcategoriaSerializer(cat.subcategorias.all(), many=True).data
            })
        return Response(result)

class SubcategoriaViewSet(viewsets.ModelViewSet):
    queryset = Subcategoria.objects.all()
    serializer_class = SubcategoriaSerializer
    permission_classes = [IsAuthenticated]

# --- VISTA PARA MANEJAR LOS EQUIPOS ---
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Equipo, Categoria, Subcategoria, Persona
from .serializers import EquipoSerializer
import pandas as pd
import io

class EquipoViewSet(viewsets.ModelViewSet):
    serializer_class = EquipoSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        # export_excel maneja su propia autenticación (token por query param para Brave)
        if self.action == 'export_excel':
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        role = getattr(getattr(user, 'persona', None), 'rol', 'VIEWER')
        
        if role in ['ADMIN', 'TECNICO']:
            queryset = Equipo.objects.all()
        else:
            # Los visualizadores solo ven equipos asignados a ellos
            # Buscamos la Persona vinculada al usuario
            persona = getattr(user, 'persona', None)
            if persona:
                queryset = Equipo.objects.filter(usuario_asignado__iexact=persona.nombre)
            else:
                queryset = Equipo.objects.none()

        subcat_id = self.request.query_params.get('subcategoria')
        cat_id = self.request.query_params.get('categoria')
        
        if subcat_id:
            queryset = queryset.filter(subcategoria_id=subcat_id)
        if cat_id:
            queryset = queryset.filter(subcategoria__categoria_id=cat_id)
            
        return queryset

    def perform_create(self, serializer):
        serializer.save(creado_por=self.request.user)


    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        import io
        from datetime import date, datetime
        from rest_framework.authtoken.models import Token as AuthToken

        # Soporte para autenticación por query param (para window.open en Brave)
        auth_token_param = request.query_params.get('auth_token')
        if auth_token_param and not request.user.is_authenticated:
            try:
                token_obj = AuthToken.objects.get(key=auth_token_param)
                request.user = token_obj.user
            except AuthToken.DoesNotExist:
                return Response({"error": "Token inválido"}, status=401)

        if not request.user.is_authenticated:
            return Response({"error": "No autenticado"}, status=401)

        try:
            role = getattr(getattr(request.user, 'persona', None), 'rol', 'VIEWER')
            if role in ['ADMIN', 'TECNICO']:
                queryset = Equipo.objects.all()
            else:
                persona = getattr(request.user, 'persona', None)
                queryset = Equipo.objects.filter(usuario_asignado__iexact=persona.nombre) if persona else Equipo.objects.none()

            estado = request.query_params.get('estado')
            cat_id = request.query_params.get('categoria')

            if estado and estado != 'ALL':
                queryset = queryset.filter(estado=estado)

            if cat_id:
                queryset = queryset.filter(subcategoria__categoria_id=cat_id)

            print(f"DEBUG: Exportando {queryset.count()} equipos (Filtros: Estado={estado}, Cat={cat_id})")

            column_mapping = {
                'subcategoria__categoria__nombre': 'CATEGORÍA',
                'subcategoria__nombre': 'SUBCATEGORÍA',
                'nombre_equipo': 'NOMBRE DEL EQUIPO',
                'serie': 'NÚMERO DE SERIE',
                'marca': 'MARCA',
                'modelo': 'MODELO',
                'activo_fijo': 'ACTIVO FIJO ID',
                'estado': 'ESTADO ACTUAL',
                'usuario_asignado': 'RESPONSABLE / USUARIO',
                'departamento': 'DEPARTAMENTO',
                'fecha_ingreso': 'FECHA INGRESO (INST.)',
            }

            if estado == 'ASIGNADO' or estado == 'ALL' or not estado:
                column_mapping['fecha_asignacion'] = 'FECHA DE ASIGNACIÓN'

            if estado == 'BAJA' or estado == 'ALL' or not estado:
                column_mapping['fecha_baja'] = 'FECHA DE BAJA'
                if estado == 'BAJA':
                    column_mapping['novedad'] = 'OBSERVACIONES / NOVEDAD DE BAJA'

            data = list(queryset.values(*column_mapping.keys()))

            # Normalizar valores: None → "", dates → string iso
            for row in data:
                for key, val in row.items():
                    if key == 'activo_fijo' and (val is None or str(val).strip() == ''):
                        row[key] = 'SIN ACTIVO FIJO'
                    elif key == 'usuario_asignado' and (val is None or str(val).strip() == ''):
                        row[key] = 'SIN RESPONSABLE'
                    elif val is None:
                        row[key] = ''
                    elif isinstance(val, (date, datetime)):
                        row[key] = val.strftime('%Y-%m-%d')

            if data:
                df = pd.DataFrame(data)
                df = df[list(column_mapping.keys())]  # garantizar orden
                df = df.sort_values(by=['subcategoria__categoria__nombre', 'subcategoria__nombre', 'nombre_equipo'])
            else:
                df = pd.DataFrame(columns=list(column_mapping.keys()))

            df = df.rename(columns=column_mapping)
            # Reemplazar NaN residuales
            df = df.fillna('')

            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # ── Hoja 1: Inventario Detallado ─────────────────────────────
                df.to_excel(writer, index=False, sheet_name='Inventario Detallado')
                workbook = writer.book
                worksheet = writer.sheets['Inventario Detallado']

                header_fill = PatternFill(start_color="1E1E2D", end_color="1E1E2D", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                cell_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin', color='4B5563'),
                    right=Side(style='thin', color='4B5563'),
                    top=Side(style='thin', color='4B5563'),
                    bottom=Side(style='thin', color='4B5563'),
                )

                # Encabezados
                for col_num, col_name in enumerate(df.columns.values, start=1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                    worksheet.column_dimensions[cell.column_letter].width = 26

                # Filas de datos
                for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.alignment = cell_align
                        cell.border = thin_border

                # ── Hoja 2: Métricas y Análisis ───────────────────────────────
                summary_sheet = workbook.create_sheet("Métricas y Análisis")
                title_font = Font(bold=True, size=12)
                hdr_font   = Font(bold=True, size=10)

                # ── Bloque: Resumen por Estado ──
                summary_sheet["A1"] = "RESUMEN POR ESTADO"
                summary_sheet["A1"].font = title_font
                summary_sheet["A2"] = "ESTADO"
                summary_sheet["B2"] = "CANTIDAD"
                summary_sheet["A2"].font = hdr_font
                summary_sheet["B2"].font = hdr_font

                if not df.empty and 'ESTADO ACTUAL' in df.columns:
                    status_counts = df['ESTADO ACTUAL'].replace('', 'Sin estado').value_counts()
                else:
                    status_counts = pd.Series(dtype=int)

                estado_start_row = 3
                for i, (sname, cnt) in enumerate(status_counts.items(), start=estado_start_row):
                    summary_sheet.cell(row=i, column=1, value=str(sname))
                    summary_sheet.cell(row=i, column=2, value=int(cnt))

                if not status_counts.empty:
                    pie = PieChart()
                    lbl_ref  = Reference(summary_sheet, min_col=1, min_row=estado_start_row,
                                         max_row=estado_start_row + len(status_counts) - 1)
                    data_ref = Reference(summary_sheet, min_col=2, min_row=2,
                                         max_row=estado_start_row + len(status_counts) - 1)
                    pie.add_data(data_ref, titles_from_data=True)
                    pie.set_categories(lbl_ref)
                    pie.title = "Distribución por Estado"
                    summary_sheet.add_chart(pie, "D2")

                # ── Bloque: Top 5 Marcas ──
                marca_block_row = estado_start_row + len(status_counts) + 3
                summary_sheet.cell(row=marca_block_row,     column=1, value="TOP 5 MARCAS").font = title_font
                summary_sheet.cell(row=marca_block_row + 1, column=1, value="MARCA").font  = hdr_font
                summary_sheet.cell(row=marca_block_row + 1, column=2, value="CANTIDAD").font = hdr_font

                if not df.empty and 'MARCA' in df.columns:
                    marca_counts = df['MARCA'].replace('', 'Sin marca').value_counts().head(5)
                else:
                    marca_counts = pd.Series(dtype=int)

                marca_data_start = marca_block_row + 2
                for i, (mname, cnt) in enumerate(marca_counts.items(), start=marca_data_start):
                    summary_sheet.cell(row=i, column=1, value=str(mname))
                    summary_sheet.cell(row=i, column=2, value=int(cnt))

                if not marca_counts.empty:
                    bar = BarChart()
                    bar.type  = "col"
                    bar.style = 10
                    bar.title = "Principales Marcas en Inventario"
                    bar.y_axis.title = "Unidades"
                    bar.x_axis.title = "Marca"
                    bar_lbl_ref  = Reference(summary_sheet, min_col=1, min_row=marca_data_start,
                                             max_row=marca_data_start + len(marca_counts) - 1)
                    bar_data_ref = Reference(summary_sheet, min_col=2, min_row=marca_block_row + 1,
                                             max_row=marca_data_start + len(marca_counts) - 1)
                    bar.add_data(bar_data_ref, titles_from_data=True)
                    bar.set_categories(bar_lbl_ref)
                    chart_anchor = f"D{marca_block_row}"
                    summary_sheet.add_chart(bar, chart_anchor)

                # ── Totales generales ──
                total_row = marca_data_start + len(marca_counts) + 3
                summary_sheet.cell(row=total_row, column=1, value="TOTAL DE EQUIPOS")
                summary_sheet.cell(row=total_row, column=1).font = Font(bold=True)
                summary_sheet.cell(row=total_row, column=2, value=len(df))

                summary_sheet.column_dimensions['A'].width = 28
                summary_sheet.column_dimensions['B'].width = 15

            output.seek(0)
            filename = f"Reporte_{estado if (estado and estado != 'ALL') else 'Global'}_Inventario_TIC.xlsx"
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No se proporcionó ningún archivo"}, status=400)

        import random
        import string
        import datetime
        from django.db import IntegrityError

        try:
            # Cargar el libro de trabajo Excel
            xls = pd.ExcelFile(file)
            
            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []
            warnings = []
            
            # Helper para validar números de serie
            def is_valid_serie(val):
                if pd.isna(val):
                    return False
                val_str = str(val).strip().upper()
                if val_str in ['', 'N/A', 'NONE', 'NULL', 'SIN SERIE', '......', '.......', '****', '............', '************', '------------']:
                    return False
                if all(c in '.-* ' for c in val_str):
                    return False
                return True

            # Helper para limpiar strings
            def clean_str(val):
                if pd.isna(val):
                    return None
                val_str = str(val).strip().upper()
                if val_str in ['', 'N/A', 'NONE', 'NULL', '......', '............', '************', '------------']:
                    return None
                return val_str

            # Helper para parsear fechas
            def parse_excel_date(val):
                if pd.isna(val) or str(val).strip() in ['', 'N/A', 'NONE', 'NULL', '......', '............']:
                    return None
                try:
                    if isinstance(val, (pd.Timestamp, datetime.date, datetime.datetime)):
                        return pd.to_datetime(val).date()
                    if isinstance(val, (int, float)):
                        return (pd.to_datetime('1899-12-30') + pd.to_timedelta(val, unit='D')).date()
                    return pd.to_datetime(str(val).strip()).date()
                except:
                    return None

            # Helper para clasificar Subcategorías
            def find_subcategoria(nombre_eq):
                if not nombre_eq:
                    return None
                nombre_lower = nombre_eq.lower()
                
                categories_mapping = [
                    {
                        "categoria": "Hardware y Componentes Internos",
                        "color": "#3b82f6",
                        "subcategorias": [
                            {
                                "nombre": "Placas y Memoria (RAM/Motherboards/NIC)",
                                "keywords": ['ram', 'memoria ram', 'motherboard', 'placa', 'tarjeta de red', 'nic']
                            },
                            {
                                "nombre": "Almacenamiento (HDD/SSD/Externos)",
                                "keywords": ['ssd', 'hdd', 'disco', 'disk', 'cd-r', 'dvd-r']
                            },
                            {
                                "nombre": "Unidades de Procesamiento (CPUs/GPUs)",
                                "keywords": ['cpu', 'gpu', 'tarjeta de video', 'procesador']
                            },
                            {
                                "nombre": "Energía (PSU/Adaptadores/Baterías)",
                                "keywords": ['bateria', 'ups', 'cargador', 'fuente', 'adaptador de voltaje', 'adaptador de energia', 'adaptador 12v', 'pila']
                            },
                            {
                                "nombre": "Carcasas y Gabinetes",
                                "keywords": ['case', 'gabinete', 'carcasa']
                            }
                        ]
                    },
                    {
                        "categoria": "Equipos de Comunicación y Telefonía",
                        "color": "#10b981",
                        "subcategorias": [
                            {
                                "nombre": "Networking (Routers/Switches/APs/Firewalls)",
                                "keywords": ['router', 'switch', 'ap', 'access point', 'firewall', 'antena', 'modem', 'nanostation', 'adaptador de red', 'transceiver']
                            },
                            {
                                "nombre": "Telefonía (IP/Analógicos/PBX)",
                                "keywords": ['telefono', 'intercomunicador', 'pbx', 'auricular', 'headset']
                            }
                        ]
                    },
                    {
                        "categoria": "Periféricos y Accesorios",
                        "color": "#f59e0b",
                        "subcategorias": [
                            {
                                "nombre": "Entrada/Salida (Teclados/Mouses/Monitores)",
                                "keywords": ['teclado', 'mouse', 'raton', 'monitor', 'pantalla', 'parlante', 'altavoz', 'microfono', 'webcam', 'camara']
                            },
                            {
                                "nombre": "Imagen (Proyectores/Scanners/Printers)",
                                "keywords": ['proyector', 'scanner', 'impresora', 'printer', 'fotocopiadora']
                            }
                        ]
                    },
                    {
                        "categoria": "Cableado y Conectividad",
                        "color": "#8b5cf6",
                        "subcategorias": [
                            {
                                "nombre": "Red (UTP/Fibra/Patch Cords)",
                                "keywords": ['cable', 'patch', 'utp', 'fibra']
                            },
                            {
                                "nombre": "Video (HDMI/DP/VGA/DVI)",
                                "keywords": ['hdmi', 'vga', 'displayport', 'dvi', 'dp']
                            },
                            {
                                "nombre": "Datos y Poder (USB/C13/Serial)",
                                "keywords": ['usb', 'c13', 'serial', 'otg']
                            }
                        ]
                    },
                    {
                        "categoria": "Insumos de Limpieza y Mantenimiento",
                        "color": "#ec4899",
                        "subcategorias": [
                            {
                                "nombre": "Gestión Térmica (Pasta/Pads)",
                                "keywords": ['pasta termica', 'pad termico', 'thermal']
                            },
                            {
                                "nombre": "Limpieza Interna (Aire/Contact Cleaner)",
                                "keywords": ['limpiador', 'aire comprimido', 'contact cleaner', 'isopropilico']
                            },
                            {
                                "nombre": "Limpieza de Superficies",
                                "keywords": ['superficies', 'alcohol', 'gel', 'paño', 'wipe']
                            }
                        ]
                    },
                    {
                        "categoria": "Lubricantes y Químicos Especializados",
                        "color": "#06b6d4",
                        "subcategorias": [
                            {
                                "nombre": "Aceites de Precisión",
                                "keywords": ['aceite']
                            },
                            {
                                "nombre": "Grasas (Litio/Grafito)",
                                "keywords": ['grasa', 'grasas']
                            },
                            {
                                "nombre": "Aflojatodo y Limpiadores de Rodillos",
                                "keywords": ['wd-40', 'aflojatodo', 'limpiador rodillos']
                            }
                        ]
                    },
                    {
                        "categoria": "Herramientas y Misceláneos",
                        "color": "#6366f1",
                        "subcategorias": [
                            {
                                "nombre": "Herramientas (Screwdrivers/Multimeters/Testers)",
                                "keywords": ['destornillador', 'multimetro', 'tester', 'pinza', 'cautin', 'herramienta']
                            },
                            {
                                "nombre": "Consumibles de Oficina (Tóner/Tinta/Papel)",
                                "keywords": ['toner', 'tinta', 'cartucho', 'papel', 'cinta']
                            }
                        ]
                    }
                ]
                
                # 1. Coincidencia por palabras clave con creación automática
                for cat_data in categories_mapping:
                    for sub_data in cat_data["subcategorias"]:
                        if any(kw in nombre_lower for kw in sub_data["keywords"]):
                            cat_obj, _ = Categoria.objects.get_or_create(
                                nombre=cat_data["categoria"],
                                defaults={"color": cat_data["color"]}
                            )
                            sub_obj, _ = Subcategoria.objects.get_or_create(
                                categoria=cat_obj,
                                nombre=sub_data["nombre"]
                            )
                            return sub_obj
                
                # 2. Fallback por coincidencia parcial de palabras
                words = [w for w in nombre_lower.split() if len(w) > 3]
                for w in words:
                    sub = Subcategoria.objects.filter(nombre__icontains=w).first()
                    if sub:
                        return sub
                
                # 3. Fallback con creación de Categoría por Defecto
                cat_other, _ = Categoria.objects.get_or_create(
                    nombre="Otros Equipos y Dispositivos",
                    defaults={"color": "#6b7280"}
                )
                sub_other, _ = Subcategoria.objects.get_or_create(
                    categoria=cat_other,
                    nombre="Otros"
                )
                return sub_other

            # Procesar todas las pestañas (sheets) del archivo Excel
            for sheet_name in xls.sheet_names:
                df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                
                # Detectar la fila de encabezados (buscando en las primeras 10 filas)
                header_row_idx = None
                for idx in range(min(10, len(df_raw))):
                    row_vals = [str(x).strip().lower() for x in df_raw.iloc[idx].tolist() if pd.notna(x)]
                    if any('serie' in x for x in row_vals) or (any('equipo' in x for x in row_vals) and any('marca' in x for x in row_vals)):
                        header_row_idx = idx
                        break
                
                if header_row_idx is None:
                    continue  # La pestaña no contiene una tabla de inventario reconocible

                # Cargar la pestaña desde la fila de encabezado detectada
                df = pd.read_excel(xls, sheet_name=sheet_name, skiprows=header_row_idx)
                
                # Limpiar nombres de columnas
                df.columns = [str(c).strip().upper().replace('\n', ' ') for c in df.columns]
                
                # Heurística de mapeo de columnas
                mapping_cols = {
                    'nombre_equipo': ['EQUIPO', 'PRODUCTO', 'MONITORES ASIGNADOS'],
                    'marca': ['MARCA'],
                    'serie': ['SERIE', 'SERIE DE EQUIPO'],
                    'modelo': ['MODELO'],
                    'activo_fijo': ['ACTIVO FIJO', 'ACTIVO'],
                    'estado_raw': ['ESTADO'],
                    'usuario_raw': ['USUARIO', 'RESPONSABLE', 'ENTREGADO A'],
                    'fecha_raw': ['FECHA', 'FECHA ENTREGA', 'FECHA INGRESO'],
                    'novedad_raw': ['NOVEDAD', 'MOTIVO', 'OBSERVACIONES']
                }
                
                # Resolver qué columna exacta corresponde en esta hoja
                col_indices = {}
                for key, alternatives in mapping_cols.items():
                    for alt in alternatives:
                        match = [c for c in df.columns if alt in c]
                        if match:
                            col_indices[key] = match[0]
                            break

                # Obligatorio tener al menos EQUIPO y SERIE en la pestaña
                if 'nombre_equipo' not in col_indices or 'serie' not in col_indices:
                    continue

                for index, row in df.iterrows():
                    try:
                        raw_eq = row[col_indices['nombre_equipo']] if 'nombre_equipo' in col_indices else None
                        raw_se = row[col_indices['serie']] if 'serie' in col_indices else None
                        
                        nombre_equipo = clean_str(raw_eq)
                        serie_clean = clean_str(raw_se)
                        
                        # 1. Validaciones básicas de la fila
                        if not nombre_equipo:
                            continue
                            
                        if not is_valid_serie(serie_clean):
                            skipped_count += 1
                            warnings.append(f"Hoja '{sheet_name}' - Fila {index + header_row_idx + 2}: Registro '{nombre_equipo}' omitido por número de serie vacío o inválido ('{raw_se}').")
                            continue

                        # Limpieza y extracción de los demás campos
                        marca = clean_str(row[col_indices['marca']]) if 'marca' in col_indices else None
                        modelo = clean_str(row[col_indices['modelo']]) if 'modelo' in col_indices else None
                        activo_fijo = clean_str(row[col_indices['activo_fijo']]) if 'activo_fijo' in col_indices else None
                        estado_raw = clean_str(row[col_indices['estado_raw']]) if 'estado_raw' in col_indices else ''
                        usuario_raw = clean_str(row[col_indices['usuario_raw']]) if 'usuario_raw' in col_indices else None
                        fecha_raw = row[col_indices['fecha_raw']] if 'fecha_raw' in col_indices else None
                        novedad_raw = clean_str(row[col_indices['novedad_raw']]) if 'novedad_raw' in col_indices else None

                        # Soporte para Sección o Bloque como Departamento
                        departamento = None
                        dept_col = [c for c in df.columns if 'SECCION' in c or 'BLOQUE' in c or 'DEPARTAMENTO' in c]
                        if dept_col:
                            departamento = clean_str(row[dept_col[0]])

                        # Limpieza de usuario si es un placeholder de celdas vacías
                        if usuario_raw and usuario_raw in ['', 'BODEGA', 'SIN ASIGNAR', 'N/A', '----------', '............', '******']:
                            usuario_raw = None

                        # Normalización de Estado
                        estado = 'DISPONIBLE'
                        if 'BAJA' in estado_raw or 'RETIRADO' in estado_raw or 'ELIMINADO' in estado_raw:
                            estado = 'BAJA'
                        elif 'REPARACION' in estado_raw or 'MANTENIMIENTO' in estado_raw or 'SERVICIO' in estado_raw:
                            estado = 'REPARACION'
                        elif 'ENTREGADO' in estado_raw or 'ASIGNADO' in estado_raw or 'ACTIVO' in estado_raw or usuario_raw:
                            estado = 'ASIGNADO'

                        # Heurística especial: Si el estado es Entregado pero no hay columna de usuario,
                        # pero en NOVEDAD viene el nombre de la persona (típico de hojas 'Entregado')
                        if estado == 'ASIGNADO' and not usuario_raw:
                            if novelty_contains_name := clean_str(novedad_raw):
                                if 'ENTREGA' in sheet_name.upper() or 'ENTREGADO' in sheet_name.upper() or 'PENDIENTES' in sheet_name.upper():
                                    if len(novelty_contains_name) < 35 and not any(x in novelty_contains_name for x in ['SOPLADA', 'DAÑADO', 'DAÑADA', 'ROTO', 'BISAGRA', 'PANTALLA', 'MALLA', 'TOTAL', 'TICKET']):
                                        usuario_raw = novelty_contains_name
                                        novedad_raw = None

                        usuario_asignado = usuario_raw

                        # 2. Agregar automáticamente a la Persona como usuario si no existe
                        if usuario_asignado:
                            persona_existente = Persona.objects.filter(nombre__iexact=usuario_asignado).first()
                            if not persona_existente:
                                # Generar identificación única aleatoria de 10 dígitos
                                identificacion = ''.join(random.choices(string.digits, k=10))
                                while Persona.objects.filter(identificacion=identificacion).exists():
                                    identificacion = ''.join(random.choices(string.digits, k=10))
                                
                                Persona.objects.create(
                                    nombre=usuario_asignado,
                                    identificacion=identificacion,
                                    rol='VIEWER'
                                )

                        # 3. Conversión y parsing de fechas
                        fecha_ing = parse_excel_date(fecha_raw)
                        fecha_baja = parse_excel_date(fecha_raw) if estado == 'BAJA' else None
                        fecha_asignacion = parse_excel_date(fecha_raw) if estado == 'ASIGNADO' else None

                        # 4. Validación de Activo Fijo (Garantizar unicidad en la BD)
                        if activo_fijo:
                            existing_af = Equipo.objects.filter(activo_fijo__iexact=activo_fijo).exclude(serie__iexact=serie_clean).first()
                            if existing_af:
                                warnings.append(f"Hoja '{sheet_name}' - Fila {index + header_row_idx + 2}: El código de activo fijo '{activo_fijo}' en el equipo '{nombre_equipo}' ya pertenece a otro equipo registrado (Serie: {existing_af.serie}). Se dejó en blanco para este registro.")
                                activo_fijo = None

                        # Clasificar dinámicamente según nombre del equipo
                        subcat = find_subcategoria(nombre_equipo)

                        # 5. Buscar si el equipo ya existe (repetido por Serie) para actualizar
                        equipo_existente = Equipo.objects.filter(serie__iexact=serie_clean).first()

                        if equipo_existente:
                            # Se repite. Aplicamos la acción de actualización correspondiente
                            # A. Reasignar si viene un usuario diferente
                            if usuario_asignado and equipo_existente.usuario_asignado != usuario_asignado:
                                old_u = equipo_existente.usuario_asignado or "BODEGA"
                                equipo_existente.usuario_asignado = usuario_asignado
                                equipo_existente.estado = 'ASIGNADO'
                                if fecha_asignacion:
                                    equipo_existente.fecha_asignacion = fecha_asignacion
                                elif not equipo_existente.fecha_asignacion:
                                    equipo_existente.fecha_asignacion = datetime.date.today()
                                equipo_existente.novedad = f"Reasignado en importación Excel de {old_u} a {usuario_asignado}"
                            
                            # B. Dar de baja si se indica
                            if estado == 'BAJA' and equipo_existente.estado != 'BAJA':
                                equipo_existente.estado = 'BAJA'
                                equipo_existente.fecha_baja = fecha_baja or datetime.date.today()
                                equipo_existente.novedad = novedad_raw or "Retirado en importación Excel"
                            elif estado == 'REPARACION' and equipo_existente.estado != 'REPARACION':
                                equipo_existente.estado = 'REPARACION'
                                equipo_existente.novedad = novedad_raw or "Enviado a mantenimiento en importación Excel"
                            elif estado == 'DISPONIBLE' and equipo_existente.estado not in ['DISPONIBLE', 'BAJA']:
                                equipo_existente.estado = 'DISPONIBLE'
                                equipo_existente.usuario_asignado = None
                                equipo_existente.fecha_asignacion = None

                            # C. Completar o actualizar otros datos más completos
                            equipo_existente.nombre_equipo = nombre_equipo
                            if subcat and not equipo_existente.subcategoria:
                                equipo_existente.subcategoria = subcat
                            if marca:
                                equipo_existente.marca = marca
                            if modelo:
                                equipo_existente.modelo = modelo
                            if activo_fijo:
                                equipo_existente.activo_fijo = activo_fijo
                            if departamento:
                                equipo_existente.departamento = departamento
                            if not equipo_existente.fecha_ingreso and fecha_ing:
                                equipo_existente.fecha_ingreso = fecha_ing

                            equipo_existente.save()
                            updated_count += 1
                        else:
                            # 6. Crear un nuevo equipo en la BD
                            Equipo.objects.create(
                                creado_por=request.user,
                                nombre_equipo=nombre_equipo,
                                serie=serie_clean,
                                marca=marca,
                                modelo=modelo,
                                activo_fijo=activo_fijo,
                                estado=estado,
                                usuario_asignado=usuario_asignado,
                                departamento=departamento,
                                novedad=novedad_raw or "Ingresado en importación Excel",
                                fecha_ingreso=fecha_ing or datetime.date.today(),
                                fecha_asignacion=fecha_asignacion,
                                fecha_baja=fecha_baja,
                                subcategoria=subcat
                            )
                            created_count += 1

                    except Exception as row_err:
                        errors.append(f"Hoja '{sheet_name}' - Fila {index + header_row_idx + 2}: {str(row_err)}")

            # Construir reporte y resumen final detallado
            msg = "Importación finalizada con éxito. "
            if created_count > 0:
                msg += f"Creados: {created_count} nuevos equipos. "
            if updated_count > 0:
                msg += f"Actualizados/Reasignados: {updated_count} equipos existentes. "
            if skipped_count > 0:
                msg += f"Omitidos por número de serie inválido: {skipped_count} filas. "

            return Response({
                "message": msg,
                "errors": errors,
                "warnings": warnings
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": f"Error crítico al procesar archivo Excel: {str(e)}"}, status=500)

# --- VISTA PARA REGISTRO DE USUARIOS ---
@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    username = request.data.get('username')
    email = request.data.get('email')
    password = request.data.get('password')
    first_name = request.data.get('first_name', '')

    if not username or not password:
        return Response({"error": "Usuario y contraseña requeridos"}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.filter(username=username).exists():
        return Response({"error": "Este usuario ya existe"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name
        )
        # Los nuevos registros están DESACTIVADOS por defecto hasta que el Admin apruebe
        user.is_active = False
        user.save()

        # Crear la Persona asociada con rol VIEWER por defecto
        # Usamos el email como identificación si no viene, o el username
        Persona.objects.create(
            user=user,
            nombre=first_name or username,
            identificacion=username[:20],
            email=email,
            rol='VIEWER'
        )

        # Crear perfil automáticamente
        Profile.objects.get_or_create(user=user)
        
        return Response({
            "message": "Registro exitoso. Tu cuenta está pendiente de aprobación por un administrador."
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['PUT', 'GET'])
@permission_classes([IsAuthenticated])
def update_profile(request):
    user = request.user
    profile, created = Profile.objects.get_or_create(user=user)

    if request.method == 'GET':
        return Response({
            "username": user.username,
            "name": user.first_name,
            "email": user.email,
            "picture": profile.picture
        })

    if request.method == 'PUT':
        name = request.data.get('name')
        email = request.data.get('email')
        picture = request.data.get('picture')

        if name is not None:
            user.first_name = name
        if email is not None:
            user.email = email
        user.save()

        if picture is not None:
            profile.picture = picture
            profile.save()

        return Response({
            "message": "Perfil actualizado",
            "user": {"username": user.username, "name": user.first_name, "email": user.email, "picture": profile.picture}
        })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')

    if not old_password or not new_password:
        return Response({"error": "Ambas contraseñas son requeridas"}, status=status.HTTP_400_BAD_REQUEST)

    if not user.check_password(old_password):
        return Response({"error": "La contraseña actual es incorrecta"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user.set_password(new_password)
        user.save()
        return Response({"message": "Contraseña actualizada correctamente"})
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)