import openpyxl

excel_path = r"C:\Users\DELL\Downloads\Inventario 2 (4) 2.xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True)

sheets_to_inspect = ['Buscador', 'Inventario', 'Equipos de baja', 'INVENTARIO.']

for sheet_name in sheets_to_inspect:
    if sheet_name in wb.sheetnames:
        print(f"\n=================== SHEET: {sheet_name} ===================")
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(max_row=10, values_only=True))
        for idx, r in enumerate(rows):
            print(f"Row {idx+1}: {r}")
    else:
        print(f"Sheet {sheet_name} not found.")
